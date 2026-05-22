import os
import math
import datetime
import re
from typing import List, Dict, Optional, Tuple, Any

import pythoncom
import win32com.client as win32
from win32com.client import gencache

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

import tkinter as tk
from tkinter import ttk, messagebox, filedialog


# =========================
# OE Config (Add/modify here only)
# - template_pos는 "슬롯 기준" (MIC 4개: FLW/FLI/RCC/RRW, ACC 3개: FLX/FLY/FLZ)
# - 사용자가 입력한 센서명은 런타임에 슬롯에 순서대로 매핑됨
# =========================
HK_TEMPLATE_POS = {
    # MIC slots (4)
    "MIC1-PP1": "KX2", "MIC1-PP2": "A2",  "MIC1-Oct-PP1": "QV2", "MIC1-Oct-PP2": "EY2",
    "MIC2-PP1": "LL2", "MIC2-PP2": "O2",  "MIC2-Oct-PP1": "RJ2", "MIC2-Oct-PP2": "FM2",
    "MIC3-PP1": "LZ2", "MIC3-PP2": "AC2", "MIC3-Oct-PP1": "RX2", "MIC3-Oct-PP2": "GA2",
    "MIC4-PP1": "MN2", "MIC4-PP2": "AQ2", "MIC4-Oct-PP1": "SL2", "MIC4-Oct-PP2": "GO2",
    "EXTERNAL-PP1": "QH2", "EXTERNAL-PP2": "EK2", "EXTERNAL-Oct-PP1": "WF2", "EXTERNAL-Oct-PP2": "KI2",

    # ACC slots (3)
    "ACC1-PP1": "NB2", "ACC1-PP2": "BE2", "ACC1-Oct-PP1": "SZ2", "ACC1-Oct-PP2": "HC2",
    "ACC2-PP1": "NP2", "ACC2-PP2": "BS2", "ACC2-Oct-PP1": "TN2", "ACC2-Oct-PP2": "HQ2",
    "ACC3-PP1": "OD2", "ACC3-PP2": "CG2", "ACC3-Oct-PP1": "UB2", "ACC3-Oct-PP2": "IE2",
    "ACC4-PP1": "OR2", "ACC4-PP2": "CU2", "ACC4-Oct-PP1": "UP2", "ACC4-Oct-PP2": "IS2",
    "ACC5-PP1": "PF2", "ACC5-PP2": "DI2", "ACC5-Oct-PP1": "VD2", "ACC5-Oct-PP2": "JG2",
    "ACC6-PP1": "PT2", "ACC6-PP2": "DW2", "ACC6-Oct-PP1": "VR2", "ACC6-Oct-PP2": "JU2",
}

# TODO: 실제 FAW-VW&AUDI / BYD 템플릿의 셀 매핑표가 있으면 여기만 바꾸면 됩니다.
FAWVW_AUDI_TEMPLATE_POS = dict(HK_TEMPLATE_POS)  # placeholder
BYD_TEMPLATE_POS = dict(HK_TEMPLATE_POS)         # placeholder

OE_CONFIGS = {
    "HK/HKMC": {
        "template_pos": HK_TEMPLATE_POS,
        "sheet_mode": "SURFACE",        # RR / SR / CR / SBR
        "allowed_surfaces": {"RR", "SR", "CR", "SBR"},
    },
    "FAW-VW & AUDI": {
        "template_pos": FAWVW_AUDI_TEMPLATE_POS,
        "sheet_mode": "SURFACE_SPEED",  # RR60 / SR80 ...
        "allowed_surfaces": {"RR", "SR", "CR", "SBR"},
    },
    "BYD": {
        "template_pos": BYD_TEMPLATE_POS,
        "sheet_mode": "SURFACE_SPEED",
        "allowed_surfaces": {"RR", "SR", "CR", "SBR"},
    },
}
DEFAULT_OE = "HK/HKMC"

# Template slot definitions (do not change unless templates change)
MIC_SLOT_KEYS = ["MIC1", "MIC2", "MIC3", "MIC4", "EXTERNAL"]
ACC_SLOT_KEYS = ["ACC1", "ACC2", "ACC3", "ACC4", "ACC5", "ACC6"]


def build_runtime_template_pos(base_pos: Dict[str, str],
                               mic_sensors: List[str],
                               acc_sensors: List[str]) -> Dict[str, str]:
    """
    base_pos: OE별 슬롯 기준 TEMPLATE_POS (키: FLW/FLI/RCC/RRW/FLX/FLY/FLZ)
    mic_sensors: 사용자가 입력한 MIC 센서들 (순서 유지, 최대 4)
    acc_sensors: 사용자가 입력한 ACC 센서들 (순서 유지, 최대 3)

    반환: 사용자 센서명을 key로 갖는 TEMPLATE_POS
         예) MIC 입력이 [FLW, FRW, RLW, RRW]이면
             FRW-PP1 -> (슬롯2=FLI-PP1 위치)
    """
    out: Dict[str, str] = {}

    def map_group(user_list: List[str], slot_list: List[str]):
        for i, user_sensor in enumerate(user_list):
            if i >= len(slot_list):
                break
            slot = slot_list[i]
            for suffix in ["PP1", "PP2", "Oct-PP1", "Oct-PP2"]:
                src_key = f"{slot}-{suffix}"
                dst_key = f"{user_sensor}-{suffix}"
                if src_key in base_pos:
                    out[dst_key] = base_pos[src_key]

    map_group(mic_sensors, MIC_SLOT_KEYS)
    map_group(acc_sensors, ACC_SLOT_KEYS)
    return out


# =========================
# Logger
# =========================
class FileLogger:
    def __init__(self, log_path: str):
        self.log_path = log_path
        os.makedirs(os.path.dirname(log_path), exist_ok=True)

    def _write(self, level: str, msg: str):
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{ts}] [{level}] {msg}\n"
        try:
            with open(self.log_path, "a", encoding="utf-8") as f:
                f.write(line)
        except Exception:
            pass

    def info(self, msg: str): self._write("INFO", msg)
    def warn(self, msg: str): self._write("WARN", msg)
    def error(self, msg: str): self._write("ERROR", msg)


# =========================
# COM helpers
# =========================
def ensure_testlab_app():
    try:
        return gencache.EnsureDispatch("LMSTestLabAutomation.Application")
    except Exception:
        return win32.Dispatch("LMSTestLabAutomation.Application")


def resolve_database(testlab_app):
    db_candidate = None
    book = getattr(testlab_app, "ActiveBook", None)
    if book is not None:
        db_candidate = getattr(book, "Database", None)

    if db_candidate is None:
        proj = getattr(testlab_app, "ActiveProject", None)
        if proj is not None:
            db_candidate = getattr(proj, "Database", None)

    if db_candidate is None:
        raise RuntimeError("Database 핸들을 찾지 못했습니다. Testlab에서 프로젝트가 열린 상태인지 확인하세요.")

    return db_candidate() if callable(db_candidate) else db_candidate


def cast_to(obj, iface_name: str):
    try:
        return win32.CastTo(obj, iface_name)
    except Exception:
        return obj


def to_string_list(maybe_list_or_attrmap) -> List[str]:
    if maybe_list_or_attrmap is None:
        return []
    try:
        return [str(x) for x in list(maybe_list_or_attrmap)]
    except Exception:
        pass
    try:
        out = []
        n = int(maybe_list_or_attrmap.Count)
        try:
            for i in range(n):
                out.append(str(maybe_list_or_attrmap.Item(i)))
            return out
        except Exception:
            out.clear()
            for i in range(1, n + 1):
                out.append(str(maybe_list_or_attrmap.Item(i)))
            return out
    except Exception:
        return []


def is_xyz_sensor(name: str) -> bool:
    return name.endswith("X") or name.endswith("Y") or name.endswith("Z")


def build_item_path_etc(run_path: str, sensor: str) -> str:
    base = f"{run_path}/Fixed sampling/Stationary Free run/Sections/Map statistics/Spectrum averaged"
    if is_xyz_sensor(sensor):
        return f"{base}/AutoPower {sensor}"
    return f"{base}/AutoPower {sensor} (A)"


# =========================
# Data helpers
# =========================
def to_number(v):
    if v is None:
        return None
    if isinstance(v, (tuple, list)):
        if len(v) == 0:
            return None
        return to_number(v[0])
    try:
        return float(v)
    except Exception:
        return None


def normalize_series(arr) -> List[Optional[float]]:
    out = []
    try:
        for x in list(arr):
            out.append(to_number(x))
    except Exception:
        return []
    return out


# =========================
# Run name parsing (SPEC_MARK_SURFACE_SPEED_SEQ)
# =========================
def parse_run_fields(run_name: str) -> Dict[str, Any]:
    """
    Expected:
      SPEC_MARK_(RR|SR|CR|SBR)_000060(+) or 000100(+)_Number
    Also supports: 000060+, 000060, 000070(+)
    """
    stem = os.path.splitext(os.path.basename(str(run_name)))[0]
    parts = stem.split("_")

    spec = parts[0] if len(parts) > 0 else ""
    mark = parts[1] if len(parts) > 1 else "UNKNOWN_MARK"
    surface = parts[2] if len(parts) > 2 else "UNK"

    speed_token = parts[3] if len(parts) > 3 else ""
    speed_plus = "+" in speed_token

    digits = re.sub(r"\D", "", speed_token)  # keep only 0-9
    speed = None
    if digits:
        try:
            speed = int(digits)  # "000060" -> 60
            if speed >= 1000:
                speed = int(digits.lstrip("0") or "0")
        except Exception:
            speed = None

    seq = parts[4] if len(parts) > 4 else ""

    return {
        "stem": stem,
        "spec": spec,
        "mark": mark,
        "surface": surface,
        "speed": speed,             # 60 / 70 / 100 ...
        "speed_plus": speed_plus,   # True/False
        "seq": seq,
    }


def make_sheet_key(meta: Dict[str, Any], oe_cfg: Dict[str, Any]) -> str:
    surface = str(meta.get("surface", "UNK")).upper()
    speed = meta.get("speed", None)

    mode = oe_cfg.get("sheet_mode", "SURFACE")
    if mode == "SURFACE_SPEED":
        if speed is None:
            raise ValueError(f"SURFACE_SPEED인데 speed 파싱 실패: meta={meta}")
        return f"{surface}{int(speed)}"
    return surface



# =========================
# Sorting entries by mark order
# =========================
def sort_entries_by_mark_order(entries: List[Dict[str, Any]], mark_order: List[str]) -> List[Dict[str, Any]]:
    if not entries:
        return entries
    order_index = {m: i for i, m in enumerate(mark_order or [])}

    def key_fn(e):
        m = e.get("mark", "")
        if m in order_index:
            return (0, order_index[m])
        return (1, 10**9)

    return sorted(entries, key=key_fn)


# =========================
# PP formulas
# =========================
def mic_pp1(v: Optional[float]) -> Optional[float]:
    if v is None or v <= 0:
        return None
    return 20.0 * math.log10(v / 2.0e-5)


def mic_pp2(v: Optional[float]) -> Optional[float]:
    if v is None:
        return None
    return (v * v) / 1.506


def acc_pp1(v: Optional[float]) -> Optional[float]:
    if v is None or v <= 0:
        return None
    return 20.0 * math.log10(v / 10e-6)


def acc_pp2(v: Optional[float]) -> Optional[float]:
    if v is None:
        return None
    return (v * v) / 1.5


def apply_pp(sensor: str, pp: str, y: List[Optional[float]], mic_set: set, acc_set: set) -> List[Optional[float]]:
    """
    센서명이 템플릿 기본(FLW/FLI...)이 아니라 사용자 입력(FRW/RLW...)이어도,
    MIC/ACC 입력 리스트에 포함돼 있으면 해당 변환을 적용.
    """
    s = sensor.upper()
    if s in mic_set:
        fn = mic_pp1 if pp == "PP1" else mic_pp2
    elif s in acc_set:
        fn = acc_pp1 if pp == "PP1" else acc_pp2
    else:
        return y
    return [fn(v) for v in y]


# =========================
# Excel write helpers
# =========================
def a1_to_rowcol(a1: str) -> Tuple[int, int]:
    col_letters, row = coordinate_from_string(a1)
    return int(row), column_index_from_string(col_letters)


def write_table(ws, start_a1: str, marks: List[str], x_ref: List[Optional[float]], y_by_mark: Dict[str, List[Optional[float]]]):
    r0, c0 = a1_to_rowcol(start_a1)

    ws.cell(r0, c0, "Freq")
    for j, m in enumerate(marks, start=1):
        ws.cell(r0, c0 + j, m)

    n = len(x_ref)
    for i in range(n):
        xv = x_ref[i]
        if xv is not None:
            ws.cell(r0 + 1 + i, c0, xv)
        for j, m in enumerate(marks, start=1):
            y = y_by_mark.get(m, [])
            v = y[i] if i < len(y) else None
            if v is None:
                continue
            ws.cell(r0 + 1 + i, c0 + j, v)


def build_aligned_y_by_mark(x_ref: List[Optional[float]], entries: List[Dict[str, Any]]) -> Tuple[List[str], Dict[str, List[Optional[float]]]]:
    marks = [e["mark"] for e in entries]
    y_by_mark: Dict[str, List[Optional[float]]] = {}
    for e in entries:
        m = e["mark"]
        x_run = e["x"]
        y_run = e["y"]
        idx = {x_run[k]: y_run[k] for k in range(min(len(x_run), len(y_run)))}
        y_by_mark[m] = [idx.get(xv, None) for xv in x_ref]
    return marks, y_by_mark


# =========================
# Dialogs
# =========================
class MultiSelectDialog(tk.Toplevel):
    """
    ✅ Ctrl 없이 토글 선택:
      - 클릭: 선택/해제 토글
      - 드래그로 여러개 누적 선택 가능(토글 방식)
    """
    def __init__(self, master, title: str, items: List[str], preselect: Optional[List[str]] = None):
        super().__init__(master)
        self.title(title)
        self.geometry("650x520")
        self.resizable(True, True)

        self.items = items
        self.selected: List[str] = []
        self.result_ok = False

        frm = ttk.Frame(self, padding=10)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text=title, font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(0, 8))
        ttk.Label(frm, text="Click (Single or Multiple)", foreground="#555").pack(anchor="w", pady=(0, 10))

        boxfrm = ttk.Frame(frm)
        boxfrm.pack(fill="both", expand=True)

        self.lb = tk.Listbox(boxfrm, selectmode="multiple", activestyle="dotbox")
        self.lb.pack(side="left", fill="both", expand=True)

        sb = ttk.Scrollbar(boxfrm, orient="vertical", command=self.lb.yview)
        sb.pack(side="right", fill="y")
        self.lb.configure(yscrollcommand=sb.set)

        for it in items:
            self.lb.insert("end", it)

        # preselect
        if preselect:
            idx_map = {v: i for i, v in enumerate(items)}
            for v in preselect:
                if v in idx_map:
                    self.lb.selection_set(idx_map[v])

        # ✅ toggle selection on click (no ctrl)
        self.lb.bind("<Button-1>", self._on_click_toggle, add=False)

        btnfrm = ttk.Frame(frm)
        btnfrm.pack(fill="x", pady=(10, 0))

        ttk.Button(btnfrm, text="전체 선택", command=lambda: self.lb.selection_set(0, "end")).pack(side="left")
        ttk.Button(btnfrm, text="전체 해제", command=lambda: self.lb.selection_clear(0, "end")).pack(side="left", padx=(8, 0))

        ttk.Button(btnfrm, text="취소", command=self._cancel).pack(side="right")
        ttk.Button(btnfrm, text="확인", command=self._ok).pack(side="right", padx=(0, 8))

        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self.transient(master)
        self.grab_set()
        self.wait_window(self)

    def _on_click_toggle(self, event):
        idx = self.lb.nearest(event.y)
        if idx < 0 or idx >= self.lb.size():
            return "break"
        if self.lb.selection_includes(idx):
            self.lb.selection_clear(idx)
        else:
            self.lb.selection_set(idx)
        self.lb.activate(idx)
        return "break"

    def _ok(self):
        sel = [self.items[i] for i in self.lb.curselection()]
        if not sel:
            messagebox.showwarning("선택 필요", "하나 이상 선택해주세요.")
            return
        self.selected = sel
        self.result_ok = True
        self.destroy()

    def _cancel(self):
        self.result_ok = False
        self.destroy()


class ReorderListbox(ttk.Frame):
    def __init__(self, master, items: List[str]):
        super().__init__(master)
        self.lb = tk.Listbox(self, selectmode="browse")
        self.lb.pack(side="left", fill="both", expand=True)

        sb = ttk.Scrollbar(self, orient="vertical", command=self.lb.yview)
        sb.pack(side="left", fill="y", padx=(6, 0))
        self.lb.configure(yscrollcommand=sb.set)

        for x in items:
            self.lb.insert("end", x)

        btns = ttk.Frame(self)
        btns.pack(side="left", fill="y", padx=(10, 0))

        ttk.Button(btns, text="▲ Up", command=self._up).pack(fill="x", pady=(0, 6))
        ttk.Button(btns, text="▼ Down", command=self._down).pack(fill="x")

    def _swap(self, i, j):
        a = self.lb.get(i)
        b = self.lb.get(j)
        self.lb.delete(i); self.lb.insert(i, b)
        self.lb.delete(j); self.lb.insert(j, a)
        self.lb.selection_clear(0, "end")
        self.lb.selection_set(j)
        self.lb.activate(j)

    def _up(self):
        sel = self.lb.curselection()
        if not sel:
            return
        i = sel[0]
        if i <= 0:
            return
        self._swap(i - 1, i)

    def _down(self):
        sel = self.lb.curselection()
        if not sel:
            return
        i = sel[0]
        if i >= self.lb.size() - 1:
            return
        self._swap(i, i + 1)

    def get_order(self) -> List[str]:
        return [self.lb.get(i) for i in range(self.lb.size())]


class SectionMarkOrderDialog(tk.Toplevel):
    def __init__(self, master, section_to_marks: Dict[str, List[str]], pre_map: Optional[Dict[str, List[str]]] = None):
        super().__init__(master)
        self.title("Set Mark Order (per Section)")
        self.geometry("820x600")
        self.resizable(True, True)

        self.section_to_marks = section_to_marks
        self.pre_map = pre_map or {}
        self.result_ok = False
        self.result_map: Dict[str, List[str]] = {}

        frm = ttk.Frame(self, padding=10)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text="Section별 Mark 순서를 설정하세요", font=("Segoe UI", 12, "bold")).pack(anchor="w")
        ttk.Label(frm, text="저장 순서는 위→아래 적용", foreground="#666").pack(anchor="w", pady=(4, 10))

        nb = ttk.Notebook(frm)
        nb.pack(fill="both", expand=True)

        self._widgets: Dict[str, ReorderListbox] = {}

        for section, marks in self.section_to_marks.items():
            pre = [m for m in self.pre_map.get(section, []) if m in marks]
            if pre:
                rest = [m for m in marks if m not in pre]
                merged = pre + rest
            else:
                merged = list(reversed(marks))

            tab = ttk.Frame(nb, padding=10)
            nb.add(tab, text=section)

            ttk.Label(tab, text=f"{section}", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 6))

            w = ReorderListbox(tab, merged)
            w.pack(fill="both", expand=True)
            self._widgets[section] = w

        btnfrm = ttk.Frame(frm)
        btnfrm.pack(fill="x", pady=(10, 0))
        ttk.Button(btnfrm, text="취소", command=self._cancel).pack(side="right")
        ttk.Button(btnfrm, text="확인", command=self._ok).pack(side="right", padx=(0, 8))

        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self.transient(master)
        self.grab_set()
        self.wait_window(self)

    def _ok(self):
        out = {}
        for sec, w in self._widgets.items():
            out[sec] = w.get_order()
        self.result_map = out
        self.result_ok = True
        self.destroy()

    def _cancel(self):
        self.result_ok = False
        self.destroy()


class SectionTemplateDialog(tk.Toplevel):
    def __init__(self, master, sections: List[str], existing_map: Optional[Dict[str, str]] = None):
        super().__init__(master)
        self.title("Match Section ↔ Template")
        self.geometry("1040x560")
        self.resizable(True, True)

        self.sections = sections
        self.map: Dict[str, str] = dict(existing_map or {})
        self.result_ok = False

        frm = ttk.Frame(self, padding=10)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text="Section별 Template(.xlsm) 파일을 1:1로 지정하세요", font=("Segoe UI", 12, "bold")).pack(anchor="w")
        ttk.Label(frm, text="사용법: Section 행 우클릭 → Find… 로 Template 선택", foreground="#666").pack(anchor="w", pady=(4, 10))

        mid = ttk.Frame(frm)
        mid.pack(fill="both", expand=True)

        self.tree = ttk.Treeview(mid, columns=("section", "template"), show="headings", height=18)
        self.tree.heading("section", text="Section")
        self.tree.heading("template", text="Template path (.xlsm/.xlsx)")
        self.tree.column("section", width=260, anchor="w", stretch=False)
        self.tree.column("template", width=740, anchor="w", stretch=True)
        self.tree.pack(side="left", fill="both", expand=True)

        sb = ttk.Scrollbar(mid, orient="vertical", command=self.tree.yview)
        sb.pack(side="left", fill="y", padx=(6, 0))
        self.tree.configure(yscrollcommand=sb.set)

        for s in self.sections:
            self.tree.insert("", "end", values=(s, self.map.get(s, "")))

        self.menu = tk.Menu(self, tearoff=0)
        self.menu.add_command(label="Find…", command=self._ctx_find)
        self.menu.add_command(label="Clear", command=self._ctx_clear)

        self._ctx_target_iid: Optional[str] = None
        self.tree.bind("<Button-3>", self._on_right_click)
        self.tree.bind("<Button-2>", self._on_right_click)

        bottom = ttk.Frame(frm)
        bottom.pack(fill="x", pady=(10, 0))
        ttk.Button(bottom, text="취소", command=self._cancel).pack(side="right")
        ttk.Button(bottom, text="확인", command=self._ok).pack(side="right", padx=(0, 8))

        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self.transient(master)
        self.grab_set()
        self.wait_window(self)

    def _on_right_click(self, event):
        iid = self.tree.identify_row(event.y)
        if not iid:
            return
        self._ctx_target_iid = iid
        self.tree.selection_set(iid)
        try:
            self.menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.menu.grab_release()

    def _ctx_find(self):
        if not self._ctx_target_iid:
            return
        section = self.tree.item(self._ctx_target_iid, "values")[0]
        path = filedialog.askopenfilename(
            title=f"Select Template for [{section}]",
            filetypes=[("Excel Macro Workbook", "*.xlsm"), ("Excel Workbook", "*.xlsx"), ("All files", "*.*")]
        )
        if not path:
            return
        self.map[section] = path
        self.tree.item(self._ctx_target_iid, values=(section, path))

    def _ctx_clear(self):
        if not self._ctx_target_iid:
            return
        section = self.tree.item(self._ctx_target_iid, "values")[0]
        self.map.pop(section, None)
        self.tree.item(self._ctx_target_iid, values=(section, ""))

    def _ok(self):
        missing = [s for s in self.sections if not self.map.get(s)]
        if missing:
            messagebox.showwarning("Need mapping", f"다음 Section에 Template이 지정되지 않았습니다:\n- " + "\n- ".join(missing))
            return
        bad = [s for s in self.sections if not os.path.exists(self.map.get(s, ""))]
        if bad:
            messagebox.showerror("Missing file", f"다음 Section의 Template 파일이 존재하지 않습니다:\n- " + "\n- ".join(bad))
            return
        self.result_ok = True
        self.destroy()

    def _cancel(self):
        self.result_ok = False
        self.destroy()


# =========================
# Main App
# =========================
class MainApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("VIN Report Maker (HK/HKMC, FAW-VW, BYD)")
        self.geometry("960x760")

        self.testlab = None
        self.db = None

        self.sections_all: List[str] = []
        self.sections_selected: List[str] = []
        self.runs_selected_by_section: Dict[str, List[str]] = {}

        self.section_template_map: Dict[str, str] = {}
        self.section_mark_order_map: Dict[str, List[str]] = {}

        self._export_running = False
        self._jobs: List[Tuple[str, str]] = []
        self._job_idx = 0

        self._sensors: List[str] = []   # MIC 입력 + ACC 입력 (순서 유지)
        self._mic_set: set = set()
        self._acc_set: set = set()
        self._runtime_template_pos: Dict[str, str] = {}

        self._logger: Optional[FileLogger] = None
        self._log_path: Optional[str] = None

        # section_agg[section][sheet_key][sensor] -> pack
        self.section_agg: Dict[str, Dict[str, Dict[str, Dict[str, Any]]]] = {}

        self._build_ui()
        self._init_testlab()

    def _build_ui(self):
        frm = ttk.Frame(self, padding=12)
        frm.pack(fill="both", expand=True)

        cfgfrm = ttk.LabelFrame(frm, text="Detail", padding=10)
        cfgfrm.pack(fill="x", pady=(0, 10))

        row1 = ttk.Frame(cfgfrm)
        row1.pack(fill="x", pady=(0, 8))

        ttk.Label(row1, text="OE").pack(side="left")
        self.var_oe = tk.StringVar(value=DEFAULT_OE)
        self.cmb_oe = ttk.Combobox(row1, textvariable=self.var_oe, values=list(OE_CONFIGS.keys()),
                                  state="readonly", width=18)
        self.cmb_oe.pack(side="left", padx=(8, 18))

        row2 = ttk.Frame(cfgfrm)
        row2.pack(fill="x")

        ttk.Label(row2, text="MIC Sensors (max 5)").pack(side="left")
        self.var_mic = tk.StringVar(value="FLW,FLI,RCC,RRW,FLT")
        ttk.Entry(row2, textvariable=self.var_mic, width=40).pack(side="left", padx=(8, 18))

        ttk.Label(row2, text="ACC Sensors (max 6)").pack(side="left")
        self.var_acc = tk.StringVar(value="FLX,FLY,FLZ,RLX,RLY,RLZ")
        ttk.Entry(row2, textvariable=self.var_acc, width=30).pack(side="left", padx=(8, 0))

        selfrm = ttk.LabelFrame(frm, text="Selection", padding=10)
        selfrm.pack(fill="both", expand=True)

        row3 = ttk.Frame(selfrm)
        row3.pack(fill="x", pady=(0, 10))
        ttk.Button(row3, text="Select Sections", command=self._select_sections).pack(side="left")
        ttk.Button(row3, text="Select Runs", command=self._select_runs).pack(side="left", padx=(8, 0))
        ttk.Button(row3, text="Set Mark Order", command=self._set_mark_order_per_section).pack(side="left", padx=(18, 0))
        ttk.Button(row3, text="Match Section ↔ Templates", command=self._match_section_templates).pack(side="left", padx=(18, 0))

        mapfrm = ttk.LabelFrame(selfrm, text="Section → Template mapping", padding=10)
        mapfrm.pack(fill="x", pady=(0, 10))
        self.var_map_summary = tk.StringVar(value="(not set)")
        ttk.Label(mapfrm, textvariable=self.var_map_summary, foreground="#444").pack(anchor="w")

        treefrm = ttk.Frame(selfrm)
        treefrm.pack(fill="both", expand=True, pady=(0, 10))

        self.tree = ttk.Treeview(treefrm, show="tree")
        self.tree.pack(side="left", fill="both", expand=True)

        sb = ttk.Scrollbar(treefrm, orient="vertical", command=self.tree.yview)
        sb.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=sb.set)

        self.lbl_status = ttk.Label(selfrm, text="Ready")
        self.lbl_status.pack(anchor="w", pady=(2, 4))

        progfrm = ttk.Frame(selfrm)
        progfrm.pack(fill="x", pady=(10, 0))

        self.var_prog = tk.DoubleVar(value=0.0)
        self.progbar = ttk.Progressbar(progfrm, variable=self.var_prog, maximum=100.0)
        self.progbar.pack(side="left", fill="x", expand=True)

        self.var_prog_text = tk.StringVar(value="0%")
        ttk.Label(progfrm, textvariable=self.var_prog_text, width=6, anchor="e").pack(side="left", padx=(10, 0))

        self.var_step = tk.StringVar(value="")
        ttk.Label(selfrm, textvariable=self.var_step, foreground="#555").pack(anchor="w", pady=(6, 0))

        self.var_loghint = tk.StringVar(value="")
        ttk.Label(selfrm, textvariable=self.var_loghint, foreground="#666").pack(anchor="w", pady=(6, 0))

        btnfrm = ttk.Frame(frm)
        btnfrm.pack(fill="x", pady=(10, 0))
        ttk.Button(btnfrm, text="Export (Overwrite Templates)", command=self._export_clicked).pack(side="right")
        ttk.Button(btnfrm, text="Exit", command=self.destroy).pack(side="right", padx=(0, 8))

    def _init_testlab(self):
        try:
            pythoncom.CoInitialize()
            self.testlab = ensure_testlab_app()
            self.db = resolve_database(self.testlab)

            self.sections_all = to_string_list(self.db.SectionNames)
            if not self.sections_all:
                raise RuntimeError("Section Names가 비어있습니다. 프로젝트에 Section이 있는지 확인하세요.")

            self.lbl_status.config(text=f"Testlab connected. Sections={len(self.sections_all)}")
        except Exception as e:
            messagebox.showerror("Init failed", str(e))
            self.lbl_status.config(text="Init failed: Testlab/Project open 상태 확인 필요")

    def _refresh_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        if not self.sections_selected:
            self.tree.insert("", "end", text="(No sections selected)")
            return

        for sec in self.sections_selected:
            sec_id = self.tree.insert("", "end", text=sec, open=True)
            runs = self.runs_selected_by_section.get(sec, [])
            if not runs:
                self.tree.insert(sec_id, "end", text="(No runs selected)")
            else:
                for r in runs:
                    self.tree.insert(sec_id, "end", text=r)

    def _select_sections(self):
        if self._export_running:
            messagebox.showwarning("Busy", "Export 중에는 변경할 수 없습니다.")
            return
        dlg = MultiSelectDialog(self, "Sections 선택", self.sections_all, preselect=self.sections_selected)
        if dlg.result_ok:
            self.sections_selected = dlg.selected
            self.runs_selected_by_section = {s: self.runs_selected_by_section.get(s, []) for s in self.sections_selected}

            self.section_template_map = {s: p for s, p in self.section_template_map.items() if s in self.sections_selected}
            self.section_mark_order_map = {s: v for s, v in self.section_mark_order_map.items() if s in self.sections_selected}
            self._update_mapping_summary()
            self._refresh_tree()

    def _select_runs(self):
        if self._export_running:
            messagebox.showwarning("Busy", "Export 중에는 변경할 수 없습니다.")
            return
        if not self.sections_selected:
            messagebox.showwarning("Need", "먼저 Sections를 선택하세요.")
            return

        new_map: Dict[str, List[str]] = {}
        for sec in self.sections_selected:
            elems = to_string_list(self.db.ElementNames(sec))
            runs = []
            for rn in elems:
                try:
                    if str(self.db.ElementType(f"{sec}/{rn}")) == "Run":
                        runs.append(rn)
                except Exception:
                    pass

            pre = self.runs_selected_by_section.get(sec, [])
            dlg = MultiSelectDialog(self, f"Runs 선택 - {sec}", runs, preselect=pre)
            new_map[sec] = dlg.selected if dlg.result_ok else pre

        self.runs_selected_by_section = new_map
        self._refresh_tree()

    def _update_mapping_summary(self):
        if not self.sections_selected:
            self.var_map_summary.set("(not set)")
            return
        lines = []
        for s in self.sections_selected:
            p = self.section_template_map.get(s, "")
            lines.append(f"- {s}  →  {p if p else '(not assigned)'}")
        self.var_map_summary.set("\n".join(lines))

    def _collect_marks_per_section(self) -> Dict[str, List[str]]:
        out: Dict[str, List[str]] = {}
        for sec in self.sections_selected:
            marks = []
            seen = set()
            for run in self.runs_selected_by_section.get(sec, []):
                m = parse_run_fields(run).get("mark", "")
                if not m:
                    continue
                if m not in seen:
                    marks.append(m)
                    seen.add(m)
            out[sec] = marks
        return out

    def _set_mark_order_per_section(self):
        if self._export_running:
            messagebox.showwarning("Busy", "Export 중에는 변경할 수 없습니다.")
            return
        if not self.sections_selected:
            messagebox.showwarning("Need", "먼저 Sections를 선택하세요.")
            return

        section_to_marks = self._collect_marks_per_section()
        if not any(section_to_marks.get(s) for s in self.sections_selected):
            messagebox.showwarning("Need", "선택된 Run에서 Mark를 추출하지 못했습니다.")
            return

        dlg = SectionMarkOrderDialog(self, section_to_marks=section_to_marks, pre_map=self.section_mark_order_map)
        if dlg.result_ok:
            self.section_mark_order_map = dlg.result_map

    def _match_section_templates(self):
        if self._export_running:
            messagebox.showwarning("Busy", "Export 중에는 변경할 수 없습니다.")
            return
        if not self.sections_selected:
            messagebox.showwarning("Need", "먼저 Sections를 선택하세요.")
            return

        dlg = SectionTemplateDialog(self, self.sections_selected, existing_map=self.section_template_map)
        if dlg.result_ok:
            self.section_template_map = dict(dlg.map)
            self._update_mapping_summary()

    # -------------------------
    # Export flow
    # -------------------------
    def _get_selected_jobs(self) -> List[Tuple[str, str]]:
        out = []
        for sec in self.sections_selected:
            for r in self.runs_selected_by_section.get(sec, []):
                out.append((sec, r))
        return out

    @staticmethod
    def _parse_sensor_list(txt: str) -> List[str]:
        return [s.strip().upper() for s in (txt or "").split(",") if s.strip()]

    def _export_clicked(self):
        if self._export_running:
            messagebox.showinfo("Busy", "이미 Export 중입니다.")
            return
        if not self.sections_selected:
            messagebox.showwarning("Need", "Sections를 선택하세요.")
            return

        missing_tpl = [s for s in self.sections_selected if not self.section_template_map.get(s)]
        if missing_tpl:
            messagebox.showwarning("Need mapping", "먼저 Section ↔ Template 매칭을 완료하세요.\n- " + "\n- ".join(missing_tpl))
            return

        mic_sensors = self._parse_sensor_list(self.var_mic.get())
        acc_sensors = self._parse_sensor_list(self.var_acc.get())

        if not mic_sensors and not acc_sensors:
            messagebox.showerror("Error", "MIC/ACC Sensors가 모두 비어있습니다.")
            return
        if len(mic_sensors) > 5:
            messagebox.showerror("Error", "MIC Sensors는 최대 5개까지 입력 가능합니다.")
            return
        if len(acc_sensors) > 6:
            messagebox.showerror("Error", "ACC Sensors는 최대 6개까지 입력 가능합니다.")
            return

        # MIC/ACC 구분 (PP 변환에 사용)
        self._mic_set = set(mic_sensors)
        self._acc_set = set(acc_sensors)

        # 읽을 센서 목록(순서 유지)
        self._sensors = mic_sensors + acc_sensors

        jobs = self._get_selected_jobs()
        if not jobs:
            messagebox.showwarning("Need", "선택된 Run이 없습니다.")
            return

        # runtime TEMPLATE_POS 생성 (사용자 센서명 키로 변환)
        oe_cfg = OE_CONFIGS.get(self.var_oe.get(), OE_CONFIGS[DEFAULT_OE])
        base_pos = oe_cfg["template_pos"]
        self._runtime_template_pos = build_runtime_template_pos(base_pos, mic_sensors, acc_sensors)

        if not self._runtime_template_pos:
            messagebox.showerror("Error", "TEMPLATE_POS 매핑 생성 실패(입력 센서/템플릿 슬롯 확인 필요).")
            return

        # default mark order per section if empty
        section_to_marks = self._collect_marks_per_section()
        for sec in self.sections_selected:
            if sec not in self.section_mark_order_map or not self.section_mark_order_map.get(sec):
                self.section_mark_order_map[sec] = list(reversed(section_to_marks.get(sec, [])))

        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        any_tpl = next(iter(self.section_template_map.values()))
        log_dir = os.path.join(os.path.dirname(any_tpl), "logs")
        self._log_path = os.path.join(log_dir, f"export_section_templates_{ts}.log")
        self._logger = FileLogger(self._log_path)
        self.var_loghint.set(f"Log: {self._log_path}")

        self._jobs = jobs
        self._job_idx = 0
        self._export_running = True
        self.section_agg = {}

        self.var_prog.set(0.0)
        self.var_prog_text.set("0%")
        self.var_step.set("Starting...")
        self.after(10, self._process_next_job)

    def _ensure_section_agg_slot(self, section: str, sheet_key: str, sensor: str):
        if section not in self.section_agg:
            self.section_agg[section] = {}
        if sheet_key not in self.section_agg[section]:
            self.section_agg[section][sheet_key] = {}
        if sensor not in self.section_agg[section][sheet_key]:
            self.section_agg[section][sheet_key][sensor] = {
                "x_n_ref": None, "x_o_ref": None,
                "n_entries": [], "o_entries": []
            }

    def _process_next_job(self):
        total = len(self._jobs)
        if self._job_idx >= total:
            self.var_prog.set(100.0)
            self.var_prog_text.set("100%")
            self.var_step.set("Writing into templates (per section)...")

            try:
                self._write_all_sections_into_templates_overwrite()
                self._export_running = False
                self.var_step.set("Complete")
                messagebox.showinfo("Complete", f"All templates overwritten successfully.\n\nLog:\n{self._log_path}")
            except Exception as e:
                if self._logger:
                    self._logger.error(f"Final write failed: {e}")
                self._export_running = False
                messagebox.showerror("Failed", str(e))
            return

        sec, run = self._jobs[self._job_idx]
        i = self._job_idx + 1

        pct = (self._job_idx / total) * 100.0
        self.var_prog.set(pct)
        self.var_prog_text.set(f"{pct:.0f}%")
        self.var_step.set(f"[{i}/{total}] Reading {sec} / {run}")

        try:
            self._read_one_run_to_memory(sec, run)
        except Exception as e:
            if self._logger:
                self._logger.error(f"Job error: sec={sec} run={run} err={e}")

        pct2 = (i / total) * 100.0
        self.var_prog.set(pct2)
        self.var_prog_text.set(f"{pct2:.0f}%")

        self._job_idx += 1
        self.after(10, self._process_next_job)

    def _read_one_run_to_memory(self, section: str, run: str):
        cmd = self.testlab.cmd
        run_path = f"{section}/{run}"

        meta = parse_run_fields(run)
        mark = meta["mark"]

        oe_cfg = OE_CONFIGS.get(self.var_oe.get(), OE_CONFIGS[DEFAULT_OE])
        sheet_key = make_sheet_key(meta, oe_cfg)

        if self._logger:
            self._logger.info(
                f"DEBUG OE={self.var_oe.get()} cfg={OE_CONFIGS.get(self.var_oe.get())} run={run} meta={meta} sheet_key={sheet_key}")

        # --- read blocks ---
        narrow_blocks: Dict[str, object] = {}
        for s in self._sensors:
            full_path = build_item_path_etc(run_path, s)
            obj = self.db.GetItem(full_path)
            blk = cast_to(obj, "IBlock2")
            narrow_blocks[s] = blk

        first_sensor = next(iter(narrow_blocks.keys()))
        b0 = narrow_blocks[first_sensor]
        x_n = normalize_series(list(b0.XValues))

        y_n: Dict[str, List[Optional[float]]] = {}
        for s, blk in narrow_blocks.items():
            y_n[s] = normalize_series(list(blk.YValues))

        oct_x = None
        oct_y: Dict[str, List[Optional[float]]] = {s: [] for s in y_n.keys()}
        try:
            band_1_3 = win32.constants.BandType1_3_Octave
        except Exception:
            band_1_3 = None

        if band_1_3 is not None:
            try:
                o0 = cast_to(cmd.BLOCK_OCTAVE(b0, band_1_3), "IBlock2")
                oct_x = normalize_series(list(o0.XValues))
                for s, blk in narrow_blocks.items():
                    ob = cast_to(cmd.BLOCK_OCTAVE(blk, band_1_3), "IBlock2")
                    oct_y[s] = normalize_series(list(ob.YValues))
            except Exception:
                oct_x = None

        for s in self._sensors:
            self._ensure_section_agg_slot(section, sheet_key, s)
            slot = self.section_agg[section][sheet_key][s]

            if slot["x_n_ref"] is None:
                slot["x_n_ref"] = x_n
            if slot["x_o_ref"] is None and oct_x is not None:
                slot["x_o_ref"] = oct_x

            slot["n_entries"].append({"mark": mark, "x": x_n, "y": y_n.get(s, [])})
            if oct_x is not None:
                slot["o_entries"].append({"mark": mark, "x": oct_x, "y": oct_y.get(s, [])})

    def _write_one_section_into_template_overwrite(self, section: str, template_path: str, agg_for_section: Dict[str, Dict[str, Any]]):
        ext = os.path.splitext(template_path)[1].lower()
        keep_vba = (ext == ".xlsm")

        wb = load_workbook(template_path, keep_vba=keep_vba)

        # ✅ 사용자 센서명 기반 런타임 TEMPLATE_POS 사용
        TEMPLATE_POS = self._runtime_template_pos

        # sheet check (based on agg keys)
        missing_sheets = [sk for sk in agg_for_section.keys() if sk not in wb.sheetnames]
        if missing_sheets:
            raise RuntimeError(
                f"[{section}] Template에 다음 시트가 없습니다:\n- " + "\n- ".join(missing_sheets) + f"\n\nTemplate:\n{template_path}"
            )

        mark_order = self.section_mark_order_map.get(section, [])

        for sheet_key, by_sensor in agg_for_section.items():
            ws = wb[sheet_key]

            for sensor, pack in by_sensor.items():
                sensor_u = sensor.upper()
                x_n = pack.get("x_n_ref") or []
                x_o = pack.get("x_o_ref") or []
                n_entries = sort_entries_by_mark_order(pack.get("n_entries") or [], mark_order)
                o_entries = sort_entries_by_mark_order(pack.get("o_entries") or [], mark_order)

                if x_n and n_entries:
                    marks, y_by_mark = build_aligned_y_by_mark(x_n, n_entries)

                    key = f"{sensor_u}-PP1"
                    if key in TEMPLATE_POS:
                        y_pp1 = {m: apply_pp(sensor_u, "PP1", y_by_mark[m], self._mic_set, self._acc_set) for m in marks}
                        write_table(ws, TEMPLATE_POS[key], marks, x_n, y_pp1)

                    key = f"{sensor_u}-PP2"
                    if key in TEMPLATE_POS:
                        y_pp2 = {m: apply_pp(sensor_u, "PP2", y_by_mark[m], self._mic_set, self._acc_set) for m in marks}
                        write_table(ws, TEMPLATE_POS[key], marks, x_n, y_pp2)

                if x_o and o_entries:
                    marks, y_by_mark = build_aligned_y_by_mark(x_o, o_entries)

                    key = f"{sensor_u}-Oct-PP1"
                    if key in TEMPLATE_POS:
                        y_pp1 = {m: apply_pp(sensor_u, "PP1", y_by_mark[m], self._mic_set, self._acc_set) for m in marks}
                        write_table(ws, TEMPLATE_POS[key], marks, x_o, y_pp1)

                    key = f"{sensor_u}-Oct-PP2"
                    if key in TEMPLATE_POS:
                        y_pp2 = {m: apply_pp(sensor_u, "PP2", y_by_mark[m], self._mic_set, self._acc_set) for m in marks}
                        write_table(ws, TEMPLATE_POS[key], marks, x_o, y_pp2)

        wb.save(template_path)
        if self._logger:
            self._logger.info(f"[{section}] overwritten saved: {template_path}")

    def _write_all_sections_into_templates_overwrite(self):
        for section in self.sections_selected:
            tpl = self.section_template_map.get(section)
            if not tpl:
                raise RuntimeError(f"Section '{section}'의 Template 매칭이 비어있습니다.")
            if section not in self.section_agg:
                if self._logger:
                    self._logger.warn(f"[{section}] no aggregated data. Skip.")
                continue
            self.var_step.set(f"Writing {section} -> template")
            self._write_one_section_into_template_overwrite(section, tpl, self.section_agg[section])


def main():
    pythoncom.CoInitialize()
    MainApp().mainloop()


if __name__ == "__main__":
    main()
