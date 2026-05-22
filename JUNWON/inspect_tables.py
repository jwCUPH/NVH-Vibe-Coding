import openpyxl

path = r"RAW_EXCEL\26OE030074KF_Outdoor NVH_VIN_HMC_v003_20260326235859_SX3k 19inch 3.5S 평가.xlsm"
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

if 'Report' in wb.sheetnames:
    sheet = wb['Report']
    print("--- Table 1 (FLI) Structure ---")
    # Row 22-25 are headers
    for r in range(22, 32):
        row = [sheet.cell(row=r, column=c).value for c in range(2, 16)]
        print(f"Row {r}: {row}")
    
    print("\n--- Table 2 (RCC) Structure ---")
    for r in range(38, 48):
        row = [sheet.cell(row=r, column=c).value for c in range(2, 16)]
        print(f"Row {r}: {row}")

wb.close()
