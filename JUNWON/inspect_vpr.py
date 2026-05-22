import openpyxl

path = r"RAW_EXCEL\26OE030074KF_Outdoor NVH_VIN_HMC_v003_20260326235859_SX3k 19inch 3.5S 평가.xlsm"
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

if 'VPR' in wb.sheetnames:
    vpr = wb['VPR']
    print("\nVPR sheet rows with data:")
    for r in range(1, 100):
        row_vals = [vpr.cell(row=r, column=c).value for c in range(1, 20)]
        if any(v is not None for v in row_vals):
            print(f"Row {r}: {row_vals}")

wb.close()
