import openpyxl

path = r"RAW_EXCEL\26OE030074KF_Outdoor NVH_VIN_HMC_v003_20260326235859_SX3k 19inch 3.5S 평가.xlsm"
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

for sname in ['RR', 'SR']:
    if sname in wb.sheetnames:
        sheet = wb[sname]
        print(f"--- {sname} Sheet Scan (first 20 rows, 50 columns) ---")
        for r in range(1, 21):
            row = [sheet.cell(row=r, column=c).value for c in range(1, 51)]
            if any(v is not None for v in row):
                cols = [c for c, v in enumerate(row, 1) if v is not None]
                print(f"Row {r} cols: {cols}")
                if len(cols) > 5:
                    print(f"  Row {r} samples: {[row[c-1] for c in cols[:5]]}...")
                else:
                    print(f"  Row {r} values: {[row[c-1] for c in cols]}")

wb.close()
