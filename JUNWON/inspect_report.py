import openpyxl

path = r"RAW_EXCEL\26OE030074KF_Outdoor NVH_VIN_HMC_v003_20260326235859_SX3k 19inch 3.5S 평가.xlsm"
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

if 'Report' in wb.sheetnames:
    sheet = wb['Report']
    print("--- Report B22~O36 ---")
    for r in range(22, 37):
        row = [sheet.cell(row=r, column=c).value for c in range(2, 16)] # B to O
        print(f"Row {r}: {row}")
    
    print("\n--- Report B38~O52 ---")
    for r in range(38, 53):
        row = [sheet.cell(row=r, column=c).value for c in range(2, 16)] # B to O
        print(f"Row {r}: {row}")

wb.close()
