import openpyxl

path = r"RAW_EXCEL\26OE030074KF_Outdoor NVH_VIN_HMC_v003_20260326235859_SX3k 19inch 3.5S 평가.xlsm"
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

print("Sheets:", wb.sheetnames)

if 'DB' in wb.sheetnames:
    db = wb['DB']
    print("DB G7 (Client):", db['G7'].value)
    print("DB G3 (Project):", db['G3'].value)
    print("DB G4 (Req No):", db['G4'].value)
    print("DB G22 (Vehicle 1):", db['G22'].value)
    print("DB G23 (Vehicle 2):", db['G23'].value)
    print("DB G30 (Front Pressure):", db['G30'].value)
    print("DB G32 (Pressure Unit):", db['G32'].value)

if 'Report' in wb.sheetnames:
    report = wb['Report']
    print("\nReport B22~O36 Sample:")
    for row in report.iter_rows(min_row=22, max_row=25, min_col=2, max_col=15):
        print([cell.value for cell in row])

if 'VPR' in wb.sheetnames:
    vpr = wb['VPR']
    # Check for data areas that might correspond to charts
    # Often charts are near their data or data is in hidden columns
    print("\nVPR A1~E5 Sample:")
    for row in vpr.iter_rows(min_row=1, max_row=5, min_col=1, max_col=5):
        print([cell.value for cell in row])

wb.close()
