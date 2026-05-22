from flask import Flask, render_template, request, jsonify, send_from_directory
import os
import pandas as pd
import openpyxl
import re
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime

import os
import sys

# Get the absolute path of the current script's directory
base_dir = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__, 
            template_folder=os.path.join(base_dir, 'templates'),
            static_folder=os.path.join(base_dir, 'static'))

app.config['UPLOAD_FOLDER'] = os.path.join(base_dir, 'uploads')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(base_dir, 'vpr_reports.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Database Model
class VPRReport(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    req_no = db.Column(db.String(100))
    project = db.Column(db.String(200))
    date = db.Column(db.String(50))
    content_html = db.Column(db.Text)
    raw_data = db.Column(db.Text) # JSON string of all inputs and extracted data
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

with app.app_context():
    db.create_all()

@app.route('/save', methods=['POST'])
def save_report():
    data = request.json
    if not data:
        return jsonify({'status': 'error', 'message': '데이터가 없습니다.'})

    try:
        # Create a new report entry or update existing? Let's stick to new for now
        report = VPRReport(
            req_no=data.get('req_no'),
            project=data.get('project'),
            date=data.get('date'),
            content_html=data.get('html_content'),
            raw_data=json.dumps(data.get('raw_data'))
        )
        db.session.add(report)
        db.session.commit()
        return jsonify({'status': 'success', 'report_id': report.id})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/report/<int:report_id>', methods=['GET'])
def get_report(report_id):
    report = VPRReport.query.get_or_404(report_id)
    return jsonify({
        'content': report.content_html,
        'raw_data': json.loads(report.raw_data) if report.raw_data else None
    })

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/validate', methods=['POST'])
def validate_file():
    if 'file' not in request.files:
        return jsonify({'status': 'NG', 'message': 'No file'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'status': 'NG', 'message': 'Empty filename'})

    try:
        # Simple check if file can be opened as excel
        pd.read_excel(file, nrows=1)
        return jsonify({'status': 'OK'})
    except Exception as e:
        return jsonify({'status': 'NG', 'message': str(e)})

def excel_coord_to_indices(coord):
    """A1 -> (row=1, col=1)"""
    match = re.match(r"([A-Z]+)([0-9]+)", coord)
    if not match: return None, None
    col_str, row_str = match.groups()
    col = 0
    for char in col_str:
        col = col * 26 + (ord(char) - ord('A') + 1)
    return int(row_str), col

def get_val(sheet, coord):
    r, c = excel_coord_to_indices(coord)
    if r:
        val = sheet.cell(row=r, column=c).value
        return str(val) if val is not None else ""
    return ""

def get_min_max(sheet, range_str):
    """G39~Q39 -> 'min~max'"""
    match = re.match(r"([A-Z]+)([0-9]+)~([A-Z]+)([0-9]+)", range_str)
    if not match: return "N/A"
    sc_str, sr, ec_str, er = match.groups()
    _, sc = excel_coord_to_indices(sc_str + sr)
    _, ec = excel_coord_to_indices(ec_str + er)
    row = int(sr)
    vals = []
    for c in range(sc, ec + 1):
        v = sheet.cell(row=row, column=c).value
        if isinstance(v, (int, float)): vals.append(v)
    return f"{min(vals)}~{max(vals)}" if vals else "N/A"

import json
from flask import Response

def get_unique_vals(sheet, range_str):
    """H17~Q17 -> 'Val1, Val2' (UNIQUE)"""
    match = re.match(r"([A-Z]+)([0-9]+)~([A-Z]+)([0-9]+)", range_str)
    if not match: return "N/A"
    sc_str, sr, ec_str, er = match.groups()
    _, sc = excel_coord_to_indices(sc_str + sr)
    _, ec = excel_coord_to_indices(ec_str + er)
    row = int(sr)
    vals = []
    for c in range(sc, ec + 1):
        v = sheet.cell(row=row, column=c).value
        if v is not None and str(v).strip() != "":
            vals.append(str(v).strip())
    
    unique_vals = list(dict.fromkeys(vals)) # Preserve order, remove duplicates
    return ", ".join(unique_vals) if unique_vals else "N/A"

@app.route('/extract', methods=['POST'])
def extract_data():
    if 'files' not in request.files:
        return jsonify({'status': 'error', 'message': '파일이 없습니다.'})
    
    files = request.files.getlist('files')
    
    saved_files = []
    for i, file in enumerate(files):
        temp_path = os.path.join(app.config['UPLOAD_FOLDER'], f"temp_{i}_{file.filename}")
        file.save(temp_path)
        saved_files.append((temp_path, file.filename))

    def generate():
        results = []
        file_names = [f[1] for f in saved_files]
        total_files = len(saved_files)
        
        try:
            yield f"data: {json.dumps({'status': 'progress', 'percent': 2, 'message': '시스템 초기화 중...'})}\n\n"
            
            for i, (path, name) in enumerate(saved_files):
                file_idx = i + 1
                step_prefix = f"[{file_idx}/{total_files}]"
                base_percent = 5 + (i / total_files) * 90
                file_quota = 90 / total_files
                
                yield f"data: {json.dumps({'status': 'progress', 'percent': int(base_percent), 'message': f'{step_prefix} {name} 분석 시작...'})}\n\n"
                
                wb = None
                try:
                    wb = openpyxl.load_workbook(path, data_only=True)
                    db = wb['DB'] if 'DB' in wb.sheetnames else None
                    
                    if db:
                        yield f"data: {json.dumps({'status': 'progress', 'percent': int(base_percent + file_quota * 0.3), 'message': f'{step_prefix} 기본 및 의뢰 정보 파싱 중...'})}\n\n"
                        data = {
                            'day': file_idx,
                            'client_info': get_val(db, 'G4'),
                            'project': get_val(db, 'G3'),
                            'req_no': get_val(db, 'G2'),
                            'date': get_val(db, 'F15'),
                            'testers': f"{get_val(db, 'G13')}, {get_val(db, 'G14')}",
                            'vehicle': f"{get_val(db, 'G22')} {get_val(db, 'G23')}",
                            
                            'tire_ref': {
                                'size': get_val(db, 'G17'),
                                'pattern': get_val(db, 'G18'),
                                'brand': get_val(db, 'G19'),
                                'marking': get_val(db, 'G20')
                            },
                            'tire_sample': {
                                'size': get_unique_vals(db, 'H17~Q17'),
                                'pattern': get_unique_vals(db, 'H18~Q18'),
                                'brand': get_unique_vals(db, 'H19~Q19'),
                                'marking': get_unique_vals(db, 'H20~Q20')
                            },
                            'temp_air': get_min_max(db, 'G39~Q39'),
                            'temp_road': get_min_max(db, 'G40~Q40')
                        }
                        results.append(data)
                        yield f"data: {json.dumps({'status': 'progress', 'percent': int(base_percent + file_quota * 0.9), 'message': f'{step_prefix} 데이터 추출 완료.'})}\n\n"
                except Exception as e:
                    yield f"data: {json.dumps({'status': 'progress', 'percent': int(base_percent + file_quota), 'message': f'{step_prefix} 오류: {str(e)}'})}\n\n"
                finally:
                    if wb: wb.close()
                    try: os.remove(path)
                    except: pass
            
            if not results:
                yield f"data: {json.dumps({'status': 'error', 'message': '데이터를 찾을 수 없습니다.'})}\n\n"
            else:
                yield f"data: {json.dumps({'status': 'progress', 'percent': 100, 'message': '분석 완료!'})}\n\n"
                yield f"data: {json.dumps({'status': 'success', 'data': results, 'file_names': file_names})}\n\n"
        except Exception as e:
            yield f"data: {json.dumps({'status': 'error', 'message': str(e)})}\n\n"

    return Response(generate(), mimetype='text/event-stream')

if __name__ == '__main__':
    app.run(debug=True, port=5000)
